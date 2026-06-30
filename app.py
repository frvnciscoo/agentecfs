import streamlit as st
import pandas as pd
import os
import sys
from dbfread import DBF
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import numpy as np
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import datetime
import tempfile
from io import BytesIO
import base64
import re

# --- FUNCIÓN AUXILIAR RUTAS ---
def resolver_ruta(ruta_relativa):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, ruta_relativa)

def obtener_entregas_excluidas(rutas_historicas):
    """
    Lee archivos de remates anteriores para identificar qué Entregas/Contratos 
    ya fueron procesados. Divide las entregas combinadas (ej: 'A / B').
    """
    excluidas = set()
    
    if not rutas_historicas:
        return excluidas
        
    if isinstance(rutas_historicas, str):
        rutas_historicas = [rutas_historicas]
        
    st.info(f"Analizando {len(rutas_historicas)} archivos históricos...")
    
    for ruta in rutas_historicas:
        try:
            df = pd.read_excel(ruta, header=6)
            
            col_obj = None
            if 'Entrega' in df.columns:
                col_obj = 'Entrega'
            elif 'Contrato' in df.columns:
                col_obj = 'Contrato'
            
            if col_obj:
                serie = df[col_obj].astype(str)
                
                for valor in serie:
                    partes = valor.split('/')
                    for parte in partes:
                        limpio = parte.strip().replace('.0', '')
                        if limpio:
                            excluidas.add(limpio)
                            
                st.success(f"{os.path.basename(ruta)}: Procesado correctamente.")
            else:
                st.warning(f"{os.path.basename(ruta)}: No se encontró columna 'Entrega' o 'Contrato'.")
                
        except Exception as e:
            st.error(f"Error leyendo histórico {os.path.basename(ruta)}: {e}")
            
    st.info(f"Total entregas únicas a excluir: {len(excluidas)}")
    return excluidas

def obtener_entregas_excluidas_hojas(rutas_historicas):
    """
    Lee los nombres de las HOJAS de los archivos históricos.
    En Celulosa, cada hoja es una Entrega ya procesada.
    """
    excluidas = set()
    
    if not rutas_historicas:
        return excluidas
        
    if isinstance(rutas_historicas, str):
        rutas_historicas = [rutas_historicas]
        
    st.info(f"Analizando pestañas de {len(rutas_historicas)} archivos históricos...")
    
    for ruta in rutas_historicas:
        try:
            wb = load_workbook(ruta, read_only=True, keep_links=False)
            
            for sheet_name in wb.sheetnames:
                limpio = sheet_name.strip()
                if limpio:
                    excluidas.add(limpio)
            
            st.success(f"{os.path.basename(ruta)}: Pestañas extraídas.")
            wb.close()
                
        except Exception as e:
            st.error(f"Error leyendo histórico {os.path.basename(ruta)}: {e}")
            
    st.info(f"Total entregas (hojas) a excluir: {len(excluidas)}")
    return excluidas

# ==========================================
#   NUEVA FUNCIÓN AUXILIAR DE FORMATO
# ==========================================
def agregar_cabecera_arauco(ws, datos):
    """
    Dibuja la cabecera estilo Arauco en la hoja activa (ws).
    """
    fuente_negrita = Font(bold=True, name='Calibri', size=11)
    borde_fino = Side(border_style="thin", color="000000")
    caja = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    alineacion_izq = Alignment(horizontal="left", vertical="center")
    alineacion_centro = Alignment(horizontal="center", vertical="center")

    # Fila 1
    ws['A1'] = "Nave";          ws['B1'] = datos['nave']
    ws['D1'] = "Exportador";    ws['E1'] = datos['exportador']
    
    # Fila 2
    ws['A2'] = "Destino";       ws['B2'] = datos['destino']
    ws['D2'] = "Embarcador";    ws['E2'] = datos['embarcador']
    
    # Fila 3
    ws['A3'] = "Reserva";       ws['B3'] = datos['reserva']
    ws['D3'] = "Carga";         ws['E3'] = datos['carga']
    
    # Fila 4
    ws['A4'] = "Contrato";      ws['B4'] = datos['contrato']
    ws['D4'] = "Tipo/Linea";    ws['E4'] = datos['linea']

    ws.merge_cells('B4:C4')

    for row in range(1, 5):
        celda_tit1 = ws.cell(row=row, column=1)
        celda_tit2 = ws.cell(row=row, column=4)
        
        celda_tit1.font = fuente_negrita
        celda_tit2.font = fuente_negrita
        
        celda_val1 = ws.cell(row=row, column=2)
        celda_val2 = ws.cell(row=row, column=5)
        
        for col in range(1, 6):
            celda = ws.cell(row=row, column=col)
            celda.border = caja
            celda.alignment = alineacion_izq

    ws['B4'].alignment = alineacion_centro
# ==========================================
#   PREPROCESAMIENTO ROBUSTO (TIPO CORREO)
# ==========================================
def limpiar_dataframe(df):
    """
    Aplica limpieza profunda al DataFrame para evitar errores en cruces (merges).
    1. Elimina columnas duplicadas.
    2. Elimina caracteres nulos (\x00) y espacios en todo el DataFrame.
    """
    if df.empty:
        return df
        
    # 1. Eliminar columnas duplicadas que causan colisiones
    df = df.loc[:, ~df.columns.duplicated()].copy()
    
    # 2. Limpieza profunda de strings a nivel global
    # Usamos try-except para soportar tanto versiones nuevas (map) como antiguas (applymap) de Pandas
    try:
        df = df.map(lambda x: str(x).replace('\x00', '').strip() if isinstance(x, str) else x)
    except AttributeError:
        df = df.applymap(lambda x: str(x).replace('\x00', '').strip() if isinstance(x, str) else x)
        
    return df

def leer_y_limpiar_excel(ruta, **kwargs):
    """
    Wrapper para leer Excel y limpiarlo inmediatamente.
    """
    df = pd.read_excel(ruta, **kwargs)
    return limpiar_dataframe(df)
# ==========================================
#      LÓGICA DE MADERA (NUEVA VERSIÓN TOOLS)
# ==========================================
def procesar_madera(rutas):
    """
    1. Filtra Programa contra históricos y saldos.
    2. Carga Tools unificado (reemplaza despacho, detalle, informe).
    3. Cruza con ZOOPP para clase de mercancía.
    4. Genera Remates y Pickings utilizando la nueva estructura.
    """
    st.info("Iniciando procesamiento de Madera...")
    
    def separar_entregas_multiples(df, col_entrega):
        if col_entrega not in df.columns:
            return df
        
        df[col_entrega] = df[col_entrega].astype(str)
        df[col_entrega] = df[col_entrega].str.split('/')
        df = df.explode(col_entrega)
        df[col_entrega] = df[col_entrega].str.strip().str.replace(r'\.0$', '', regex=True)
        return df

    try:
        # 1. Cargar PROGRAMA
        programa = leer_y_limpiar_excel(rutas['programa'])
        programa = separar_entregas_multiples(programa, "Entrega")
        
        if 'historico' in rutas and rutas['historico']:
            excluidas = obtener_entregas_excluidas(rutas['historico'])
            if excluidas:
                st.info(f"Filtrando {len(excluidas)} entregas históricas...")
                programa['Entrega_Str'] = programa['Entrega'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                programa = programa[~programa['Entrega_Str'].isin(excluidas)].copy()
                programa = programa.drop(columns=['Entrega_Str'])
                if programa.empty:
                    return False, "Todas las entregas del programa ya fueron procesadas en los históricos adjuntos.", []

        # 2. Cargar SALDOS
        if 'saldos' in rutas and rutas['saldos']:
            try:
                saldos = leer_y_limpiar_excel(rutas['saldos'])
                saldos = separar_entregas_multiples(saldos, "Entrega")
                st.success("Archivo Saldos cargado y normalizado.")
            except Exception as e:
                st.warning(f"Error leyendo Saldos: {e}. Continuando sin él.")
                saldos = pd.DataFrame(columns=["Entrega", "Box Saldo"])
        else:
            saldos = pd.DataFrame(columns=["Entrega", "Box Saldo"])

        saldos['Box Saldo'] = pd.to_numeric(saldos['Box Saldo'], errors='coerce')
        entregas_con_saldo = saldos.loc[saldos["Box Saldo"] != 0, "Entrega"].unique()
        
        # Filtro de Programa
        prog_filtrado = programa[
            (~programa["Entrega"].isin(entregas_con_saldo)) & 
            (programa["PRODINFO"].isin(["M.ASER.VERDE", "M.ASER. SECA", "M&B/SHOP","CLEARS","MDF MOLDURAS","MOLDURAS","BLANKS","SHOP","MOULDING&BETTER","M.PALL.SECA","M.PALL.VERDE","BASAS","AGLOMERADOS","MDF PANEL","PLYWOOD","TRUPAN","TABLERO","OSB","CHAPAS"]))
        ].copy()

        columnas_prog = ["Entrega", "Nave", "PRODINFO", "RESERVA", "DESTINO"]
        prog_filtrado = prog_filtrado[columnas_prog]
        
        try:
            nave_header = prog_filtrado["Nave"].dropna().iloc[0]
        except:
            nave_header = "SIN NAVE"

        # 3. Cargar ARCHIVO TOOLS ÚNICO
        tools = leer_y_limpiar_excel(rutas['tools'])
        
        # Identificar la columna que actúa como Entrega/Contrato en tools
        col_entrega_tools = 'Contrato' if 'Contrato' in tools.columns else 'Orden_Pedido'
        tools = separar_entregas_multiples(tools, col_entrega_tools)

        # Construir Contenedor
        def construir_contenedor(row):
            sigla = str(row.get('Cnt_Sigla', '')).strip()
            numero = str(row.get('Cnt_Nro', '')).split('.')[0].strip().zfill(6)
            dv = str(row.get('Cnt_DV', '')).strip()
            return f"{sigla}-{numero}-{dv}"

        tools['CONTENEDOR'] = tools.apply(construir_contenedor, axis=1)

        # 4. Cargar ZOOPP
        ruta_zoopp = rutas['zoopp']
        if ruta_zoopp.lower().endswith('.dbf'):
            try:
                table = DBF(ruta_zoopp, encoding='latin-1', char_decode_errors='ignore')
                zoopp = pd.DataFrame(iter(table))
                zoopp.columns = [c.lower() for c in zoopp.columns]
                zoopp = limpiar_dataframe(zoopp)
                mapeo_dbf = {
                    "loteof": "loteof,C,10", "vollote": "vollote,C,15",
                    "posped": "posped,N,6,0", "desmat": "desmat,C,40"
                }
                zoopp = zoopp.rename(columns=mapeo_dbf)
            except Exception as e:
                st.error(f"Error leyendo DBF: {e}")
                raise e
        else:
            zoopp = leer_y_limpiar_excel(rutas['zoopp'])

        zoopp['loteof,C,10'] = zoopp['loteof,C,10'].astype(str).str.strip()
        zoopp = zoopp.drop_duplicates(subset=['loteof,C,10'])

        # --- CRUCE Y PREPARACIÓN DE DATOS ---
        tools['Codigo_Barra'] = tools['Codigo_Barra'].astype(str).str.strip()
        
        # Añadir clase mercancía de ZOOPP
        tools = tools.merge(
            zoopp[['loteof,C,10', 'clase_merc']],
            left_on='Codigo_Barra', right_on='loteof,C,10', how='left'
        )

        prog_filtrado['PRODINFO'] = prog_filtrado['PRODINFO'].astype(str).str.strip()

        # Cruce Maestro: Tools + Programa
        resultado_final = tools.merge(
            prog_filtrado,
            left_on=[col_entrega_tools, 'clase_merc'],
            right_on=['Entrega', 'PRODINFO'],
            how='inner'
        )

        # Conversión de numéricos
        resultado_final["Peso_lote"] = pd.to_numeric(resultado_final["Peso_lote"].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
        resultado_final["Volumen_Lote"] = pd.to_numeric(resultado_final["Volumen_Lote"].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
        resultado_final["Tara"] = pd.to_numeric(resultado_final["Tara"].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
        resultado_final["MAXGROSS"] = pd.to_numeric(resultado_final.get("Max_Gross", 999999), errors='coerce').fillna(999999)

        # Drop duplicates por lote (similar a la lógica original)
        resultado_final = resultado_final.drop_duplicates(
            subset=["Codigo_Barra", "Entrega", "CONTENEDOR"], 
            keep="first"
        )

        # =========================================================================
        # --- GENERAR REMATE NORMAL
        # =========================================================================
        remate = (
            resultado_final.groupby(["CONTENEDOR", "Entrega", "PRODINFO"]).agg({
                "RESERVA": "first",
                "DESTINO": "first",
                "Sello_linea": "first",
                "Codigo_Barra": "count",       
                "Peso_lote": "sum",             
                "Volumen_Lote": "sum",
                "Tara": "first", 
                "MAXGROSS": "first"                         
            }).reset_index()
        )
        
        remate["PESO_BRUTO_TOTAL"] = remate["Peso_lote"] + remate["Tara"]
        contenedores_con_sobrepeso = set(
            remate[remate["PESO_BRUTO_TOTAL"] >= remate["MAXGROSS"]]["CONTENEDOR"].unique()
        )        

        remate = remate.rename(columns={
            "CONTENEDOR": "Contenedor",
            "RESERVA": "Reserva",
            "DESTINO": "Pto Destino",
            "PRODINFO": "Clase de Producto",
            "Sello_linea": "Sello",
            "Codigo_Barra": "Cantidad de Lotes",
            "Peso_lote": "Peso Total (kg)",
            "Volumen_Lote": "Volumen Total (m3)"
        })

        remate = remate[[
            "Entrega", "Reserva", "Pto Destino", "Clase de Producto",
            "Contenedor", "Sello", "Cantidad de Lotes", 
            "Peso Total (kg)", "Volumen Total (m3)"
        ]]
        remate = remate.sort_values(by=["Entrega", "Contenedor"])
        
        # Formateo visual del Remate Normal
        output = BytesIO()
        remate.to_excel(output, index=False, startrow=6, engine='openpyxl')
        
        wb = load_workbook(output)
        ws = wb.active
        fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y")
        
        ws['A1'] = "INFORME DE CONTENEDORES CONSOLIDADOS PARA EMBARQUE"
        ws['A3'] = "SAN VICENTE TERMINAL INTERNACIONAL"
        ws['A4'] = f"FECHA: {fecha_hoy}"
        ws['A5'] = f"NAVE: {nave_header}"
        
        bold_font = Font(bold=True)
        for cell in ['A1', 'A3', 'A5']:
            ws[cell].font = bold_font
            ws[cell].alignment = Alignment(horizontal="left")
        ws['A4'].alignment = Alignment(horizontal="left")

        # Merge de celdas y coloreo de sobrepeso
        columnas_a_fusionar = [1, 2, 3]  
        start_row = 8 
        
        if ws.max_row >= start_row:
            current_value = ws.cell(row=start_row, column=1).value
            merge_start = start_row

            for row in range(start_row + 1, ws.max_row + 1):
                value = ws.cell(row=row, column=1).value
                if value != current_value:
                    if merge_start != row - 1:
                        for col in columnas_a_fusionar:
                            ws.merge_cells(start_row=merge_start, start_column=col, end_row=row - 1, end_column=col)
                    current_value = value
                    merge_start = row
            if merge_start != ws.max_row:
                for col in columnas_a_fusionar:
                    ws.merge_cells(start_row=merge_start, start_column=col, end_row=ws.max_row, end_column=col)

        rojo_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        
        for row in ws.iter_rows(min_row=7):
            celda_contenedor = row[4]
            valor_contenedor = str(celda_contenedor.value).strip()
            
            if valor_contenedor in contenedores_con_sobrepeso:
                celda_contenedor.fill = rojo_fill
                
            for cell in row:
                cell.alignment = alignment_center
        
        remate_output = BytesIO()
        wb.save(remate_output)
        remate_output.seek(0)
        
        # =========================================================================
        # --- GENERAR REMATE SAG 
        # =========================================================================
        remate_sag = (
            resultado_final.groupby(["CONTENEDOR", "Entrega"]).agg({
                "RESERVA": "first",
                "DESTINO": "first",
                "PRODINFO": "first",
                "Sello_linea": "first",
                "Sello_Inspector": "first",
                "Codigo_Barra": "count",
                "Peso_lote": "sum",
                "Volumen_Lote": "sum"
            }).reset_index()
        )

        remate_sag = remate_sag.rename(columns={
            "CONTENEDOR": "Contenedor",
            "RESERVA": "Reserva",
            "DESTINO": "Pto Destino",
            "PRODINFO": "Clase de Producto",
            "Sello_linea": "Sello",
            "Sello_Inspector": "Sello Inspector", # <--- 1. AGREGAR ESTA LÍNEA
            "Codigo_Barra": "Cantidad de Lotes",
            "Peso_lote": "Peso Total (kg)",
            "Volumen_Lote": "Volumen Total (m3)"
        })

        remate_sag = remate_sag[
            ["Entrega", "Reserva", "Pto Destino", "Clase de Producto",
             "Contenedor", "Sello", "Sello Inspector", # <--- 2. QUITAR EL GUION BAJO AQUÍ
             "Cantidad de Lotes", "Peso Total (kg)", "Volumen Total (m3)"]
        ]

        remate_sag = remate_sag.sort_values(by=["Entrega", "Contenedor"])
        remate_sag_output = BytesIO()
        remate_sag.to_excel(remate_sag_output, index=False, engine='openpyxl')
        remate_sag_output.seek(0)
       
        # =========================================================================
        # --- GENERAR PICKING ORIGINAL ---
        # =========================================================================
        resultado_final['Fecha_Despacho'] = pd.to_datetime(resultado_final['Fecha_Despacho'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')

        picking_cabecera = (
            resultado_final.groupby(["CONTENEDOR", "Entrega"]).agg({
                "Sello_linea": "first",
                "RESERVA": "first",
                "Orden_Embarque": "first",
                "Peso_lote": "sum",
                "Tara": "first",
                "Fecha_Despacho": "first"
            }).reset_index()
        )

        picking_cabecera = picking_cabecera.rename(columns={
            "Sello_linea": "Sello", "RESERVA": "Reserva", "Orden_Embarque": "DUS",
            "Peso_lote": "Peso Bruto (kg)", "Tara": "Tara (kg)", "Fecha_Despacho": "Fecha Contable"
        })
        picking_cabecera["Peso Total (kg)"] = picking_cabecera["Peso Bruto (kg)"] + picking_cabecera["Tara (kg)"]
        
        vals_fijos = {
            "TPLST": "ZTPC", "Un Med Peso": "KG", "Un Med Tara": "KG", "Material Embalaje": "HC40",
            "Clase Med Transporte": "Z100", "Clave Flete": "0001", "Tipo Flete": "01",
            "Nombre Despachador": "A", "Rut Despachador": "1", "Nombre Chofer": "A",
            "RUT Chofer": "1", "Patente": "A", "Transportista": "50025", "Guia": "1", "Almacen Destino": "7004"
        }
        for k, v in vals_fijos.items():
            picking_cabecera[k] = v
            
        picking_cabecera["ID Cabecera"] = range(1, len(picking_cabecera) + 1)
        picking_cabecera = picking_cabecera.rename(columns={
            "CONTENEDOR": "ID Contenedor", "Tara (kg)": "Tara Contenedor", "Sello": "Sello Cont Nro",
            "Reserva": "Booking Nro", "Peso Bruto (kg)": "Peso Bruto Carga",
            "Peso Total (kg)": "Entrega Peso Total", "DUS": "DUS Nro"
        })
        
        cols_pick = ["ID Cabecera", "Entrega","Almacen Destino", "Fecha Contable", "Guia", "Transportista", "Patente",
                     "RUT Chofer", "Nombre Chofer", "Rut Despachador", "Nombre Despachador", "Tipo Flete",
                     "Clave Flete", "Clase Med Transporte", "Material Embalaje", "ID Contenedor", "Tara Contenedor",
                     "Un Med Tara", "Sello Cont Nro", "Booking Nro", "Peso Bruto Carga", "Un Med Peso", "DUS Nro",
                     "Entrega Peso Total", "TPLST"]
        picking_cabecera = picking_cabecera[cols_pick]

        # Tabla POSICION (Original)
        posicion = resultado_final.merge(
            picking_cabecera[['ID Cabecera', 'ID Contenedor', 'Entrega']],
            left_on=['CONTENEDOR', 'Entrega'],
            right_on=['ID Contenedor', 'Entrega'],
            how='left'
        )
        posicion['Cantidad'] = posicion.groupby(['ID Cabecera', 'Codigo_Barra'])['Codigo_Barra'].transform('count')
        posicion['ID Posicion'] = posicion.groupby('ID Cabecera')['Codigo_Barra'].rank(method='dense').astype(int)
        
        posicion = posicion[['ID Cabecera', 'ID Posicion', 'Codigo_Barra', 'Cantidad', 'Peso_lote']]
        posicion = posicion.rename(columns={'Codigo_Barra': 'Lote', 'Peso_lote': 'Peso'})
        posicion['Unidad'] = "PQT"
        posicion = posicion[['ID Cabecera', 'ID Posicion', 'Lote', 'Cantidad', 'Unidad', 'Peso']]
        posicion = posicion.sort_values(by=["ID Cabecera", "ID Posicion"]).reset_index(drop=True)

        picking_output = BytesIO()
        with pd.ExcelWriter(picking_output, engine="openpyxl") as writer:
            picking_cabecera.to_excel(writer, sheet_name="Cabecera", index=False)
            posicion.to_excel(writer, sheet_name="Posicion", index=False)
        picking_output.seek(0)

        # =========================================================================
        # --- GENERAR PICKING NUEVO ---
        # =========================================================================
        picking_cabecera_nuevo = (
            resultado_final.groupby(["CONTENEDOR", "Entrega"]).agg({
                "Sello_linea": "first",
                "RESERVA": "first",
                "Orden_Embarque": "first",
                "Peso_lote": "sum",
                "Tara": "first",
                "Fecha_Despacho": "first"
            }).reset_index()
        )

        picking_cabecera_nuevo = picking_cabecera_nuevo.rename(columns={
            "CONTENEDOR": "ID Contenedor",
            "Sello_linea": "Sello Cont Nro", 
            "RESERVA": "Booking Nro", 
            "Orden_Embarque": "DUS Nro",
            "Peso_lote": "Peso Bruto Carga", 
            "Tara": "Tara Contenedor", 
            "Fecha_Despacho": "Fecha Contable"
        })
        
        picking_cabecera_nuevo["Peso Total"] = picking_cabecera_nuevo["Peso Bruto Carga"] + picking_cabecera_nuevo["Tara Contenedor"]
        picking_cabecera_nuevo["ID Cabecera"] = range(1, len(picking_cabecera_nuevo) + 1)
        
        vals_fijos_nuevo = {
            "Centro Origen": "TD06",
            "Almacen Origen": "0100",
            "Centro Destino": "TD06",
            "Almacen Destino": "7004",
            "Guia": "1",
            "Transportista": "50025",
            "Patente": "A",
            "RUT Chofer": "1",
            "Nombre Chofer": "A",
            "Rut Despachador": "1",
            "Nombre Despachador": "A",
            "Tipo Flete": "01",
            "Clave Flete": "0001",
            "Clase Med Transporte": "Z100",
            "Material Embalaje": "HC40",
            "Un Medida Tara": "KG",
            "Un Med Peso": "KG",
            "TPLST": "ZTPC"
        }
        for k, v in vals_fijos_nuevo.items():
            picking_cabecera_nuevo[k] = v
            
        cols_pick_nuevo = [
            "ID Cabecera", "Centro Origen", "Almacen Origen", "Centro Destino", 
            "Almacen Destino", "Fecha Contable", "Guia", "Transportista", 
            "Patente", "RUT Chofer", "Nombre Chofer", "Rut Despachador", 
            "Nombre Despachador", "Tipo Flete", "Clave Flete", "Clase Med Transporte", 
            "Material Embalaje", "ID Contenedor", "Tara Contenedor", "Un Medida Tara", 
            "Sello Cont Nro", "Booking Nro", "Peso Bruto Carga", "Un Med Peso", 
            "DUS Nro", "Entrega", "Peso Total", "TPLST"
        ]
        picking_cabecera_nuevo = picking_cabecera_nuevo[cols_pick_nuevo]

        # Tabla POSICION (Nuevo)
        posicion_nuevo = resultado_final.merge(
            picking_cabecera_nuevo[['ID Cabecera', 'ID Contenedor', 'Entrega']],
            left_on=['CONTENEDOR', 'Entrega'],
            right_on=['ID Contenedor', 'Entrega'],
            how='left'
        )
        posicion_nuevo['Cantidad'] = posicion_nuevo.groupby(['ID Cabecera', 'Codigo_Barra'])['Codigo_Barra'].transform('count')
        posicion_nuevo['ID Posicion'] = posicion_nuevo.groupby('ID Cabecera')['Codigo_Barra'].rank(method='dense').astype(int)
        
        posicion_nuevo = posicion_nuevo.rename(columns={
            'Codigo_Barra': 'Lote', 
            'Peso_lote': 'Peso', 
            'ID Contenedor': 'BOX' 
        })
        posicion_nuevo['Unidad'] = "PQT"
        
        posicion_nuevo = posicion_nuevo[['ID Cabecera', 'ID Posicion', 'Lote', 'Cantidad', 'Unidad', 'Peso', 'BOX']]
        posicion_nuevo = posicion_nuevo.sort_values(by=["ID Cabecera", "ID Posicion"]).reset_index(drop=True)

        picking_nuevo_output = BytesIO()
        with pd.ExcelWriter(picking_nuevo_output, engine="openpyxl") as writer:
            picking_cabecera_nuevo.to_excel(writer, sheet_name="Cabecera", index=False)
            posicion_nuevo.to_excel(writer, sheet_name="Posicion", index=False)
        picking_nuevo_output.seek(0)

        # RETORNAMOS LOS 4 ARCHIVOS EN EL ARREGLO FINAL
        return True, "Proceso completado exitosamente", [
            ("RemateMadera.xlsx", remate_output),
            ("RemateMaderaSAG.xlsx", remate_sag_output),
            ("Picking.xlsx", picking_output),
            ("Picking_Nuevo.xlsx", picking_nuevo_output)
        ]

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []

# ==========================================
#      LÓGICA UNIFICADA: CELULOSA (BKP/EKP/UKP y DP)
# ==========================================
def procesar_celulosa(rutas):
    st.info("Iniciando procesamiento de Celulosa (BKP/EKP/UKP y DP)...")
    try:
        programa = pd.read_excel(rutas['programa'])
        
        # 1. Cargar Tools (soporta múltiples archivos si suben BKP y DP por separado)
        rutas_tools = rutas['tools']
        if isinstance(rutas_tools, str):
            rutas_tools = [rutas_tools]
            
        lista_tools = []
        for ruta in rutas_tools:
            lista_tools.append(pd.read_excel(ruta))
        tools = pd.concat(lista_tools, ignore_index=True)

        # 2. Cargar Saldos
        if 'saldos' in rutas and rutas['saldos']:
            try:
                saldos = pd.read_excel(rutas['saldos'])
            except:
                saldos = pd.DataFrame(columns=["Entrega", "Box Saldo"])
        else:
            saldos = pd.DataFrame(columns=["Entrega", "Box Saldo"])

        saldos['Entrega'] = saldos['Entrega'].astype(str).str.strip()
        programa['Entrega'] = programa['Entrega'].astype(str).str.strip()
        saldos['Box Saldo'] = pd.to_numeric(saldos['Box Saldo'], errors='coerce')
        
        # 3. Filtrar Históricos
        if 'historico' in rutas and rutas['historico']:
            excluidas = obtener_entregas_excluidas_hojas(rutas['historico'])
            if excluidas:
                st.info(f"Filtrando contra {len(excluidas)} entregas históricas (Pestañas)...")
                programa = programa[~programa['Entrega'].isin(excluidas)].copy()
                
                if programa.empty:
                    return False, "Todas las entregas del programa ya existen como hojas en los históricos adjuntos.", []

        entregas_con_saldo = saldos.loc[saldos["Box Saldo"] != 0, "Entrega"].unique()
        
        # 4. Filtrar Programa solo para Celulosa válida
        prog_filtrado = programa[
            (~programa["Entrega"].isin(entregas_con_saldo)) & 
            (programa["PRODINFO"].isin(["CEL BKP", "CEL UKP", "CEL EKP", "CEL DP"]))
        ].copy()

        if prog_filtrado.empty:
            return False, "No hay entregas válidas para Celulosa (BKP/EKP/UKP o DP) tras aplicar filtros.", []

        def obtener_linea(nav):
            nav = str(nav).upper().strip()
            if "MSC" in nav: return "MSC"
            if "ONEY" in nav: return "ONE"
            if "HLL" in nav: return "HAPAG LLOYD"
            if "MAERSK" in nav or "ML" in nav: return "MAERSK"
            return nav

        prog_filtrado['NAV_CLEAN'] = prog_filtrado['NAV'].apply(obtener_linea)
        
        metadata_dict = prog_filtrado.set_index('Entrega')[
            ['Nave', 'DESTINO', 'RESERVA', 'PRODINFO', 'NAV_CLEAN']
        ].to_dict('index')

        df_final_agrupado = pd.DataFrame()

        # ==========================================
        #   SUB-PROCESO A: CELULOSA BKP / EKP / UKP
        # ==========================================
        prog_cb = prog_filtrado[prog_filtrado["PRODINFO"].isin(["CEL BKP", "CEL UKP", "CEL EKP"])]
        if not prog_cb.empty:
            if "Contrato" in tools.columns and "Expedicion" in tools.columns:
                entregas_validas_cb = prog_cb["Entrega"].unique()
                tools_cb = tools[tools["Contrato"].astype(str).isin(entregas_validas_cb)].copy()
                
                if not tools_cb.empty:
                    tools_cb["Contenedor"] = tools_cb["Contenedor"].astype(str).str.strip()
                    tools_cb["Expedicion"] = tools_cb["Expedicion"].astype(str).str.strip()
                    
                    def normalizar_box(contenedor):
                        contenedor = str(contenedor).replace('.0', '')
                        partes = contenedor.split('-')
                        if len(partes) == 3:
                            parte_media_normalizada = partes[1].zfill(6)
                            return f"{partes[0]}-{parte_media_normalizada}-{partes[2]}"
                        return contenedor

                    tools_cb["BOX"] = tools_cb["Contenedor"].apply(normalizar_box)
                    tools_cb["TARA"] = tools_cb["Tara"]
                    tools_cb["LOTE"] = tools_cb["Expedicion"]
                    tools_cb["BULTOS"] = tools_cb["Cantidad"]
                    tools_cb["SELLO"] = tools_cb["Sello_linea"]
                    tools_cb["RESERVA"] = tools_cb["Reserva"]
                    tools_cb["DUS"] = tools_cb["Orden_Embarque"]
                    tools_cb["MAX"] = tools_cb.get("Max_Gross", 999999)

                    df_agrupado_cb = (
                        tools_cb.groupby(["Contrato", "BOX", "LOTE"], as_index=False)
                          .agg({
                              "TARA": "first", "BULTOS": "sum", "SELLO": "first",
                              "RESERVA": "first", "DUS": "first", "MAX": "first"
                          })
                    )
                    df_agrupado_cb["UNI"] = df_agrupado_cb["BULTOS"] / 8
                    df_final_agrupado = pd.concat([df_final_agrupado, df_agrupado_cb], ignore_index=True)
            else:
                st.warning("El archivo Tools no contiene las columnas necesarias para BKP/EKP/UKP (Contrato, Expedicion).")

        # ==========================================
        #   SUB-PROCESO B: CELULOSA DP
        # ==========================================
        prog_dp = prog_filtrado[prog_filtrado["PRODINFO"].isin(["CEL DP"])]
        if not prog_dp.empty:
            if "Orden_Pedido" in tools.columns and "Cnt_Sigla" in tools.columns:
                entregas_validas_dp = prog_dp["Entrega"].unique()
                tools_dp = tools[tools["Orden_Pedido"].astype(str).str.strip().isin(entregas_validas_dp)].copy()
                
                if not tools_dp.empty:
                    tools_dp['Cnt_Sigla'] = tools_dp['Cnt_Sigla'].astype(str).str.strip()
                    tools_dp['Cnt_Nro'] = tools_dp['Cnt_Nro'].astype(str).str.strip()
                    tools_dp['Cnt_DV'] = tools_dp['Cnt_DV'].astype(str).str.strip()

                    def normalizar_box_tools(row):
                        sigla = str(row['Cnt_Sigla']).strip()
                        val_num = str(row['Cnt_Nro']).split('.')[0].strip() if '.' in str(row['Cnt_Nro']) else str(row['Cnt_Nro']).strip()
                        # Extraemos solo el número antes del punto decimal para el DV
                        dv = str(row['Cnt_DV']).split('.')[0].strip() if '.' in str(row['Cnt_DV']) else str(row['Cnt_DV']).strip()
                        return f"{sigla}-{val_num.zfill(6)}-{dv}"

                    tools_dp['BOX'] = tools_dp.apply(normalizar_box_tools, axis=1)
                    tools_dp["Contrato"] = tools_dp["Orden_Pedido"].astype(str).str.strip() # Homologar nombre
                    tools_dp["TARA"] = tools_dp["Tara"]
                    tools_dp["LOTE"] = tools_dp["Marca"]
                    tools_dp["SELLO"] = tools_dp["Sello_linea"]
                    tools_dp["RESERVA"] = tools_dp["Reserva"]
                    tools_dp["DUS"] = tools_dp["Orden_Embarque"]
                    tools_dp["MAX"] = tools_dp.get("Max_Gross", 999999) 

                    tools_dp = tools_dp[tools_dp["SELLO"].notna() & (tools_dp["SELLO"].astype(str).str.strip() != "")]

                    df_agrupado_dp = tools_dp.groupby(["Contrato", "BOX", "LOTE"]).agg({
                        "TARA": "first", "SELLO": "first", "RESERVA": "first",
                        "DUS": "first", "MAX": "first"
                    })
                    
                    df_agrupado_dp["UNI"] = tools_dp.groupby(["Contrato", "BOX", "LOTE"]).size()
                    df_agrupado_dp = df_agrupado_dp.reset_index()
                    df_agrupado_dp["BULTOS"] = df_agrupado_dp["UNI"] * 8
                    
                    df_final_agrupado = pd.concat([df_final_agrupado, df_agrupado_dp], ignore_index=True)
            else:
                st.warning("El archivo Tools no contiene las columnas necesarias para DP (Orden_Pedido, Cnt_Sigla).")

        # ==========================================
        #   GENERAR EXCEL FINAL UNIFICADO
        # ==========================================
        if df_final_agrupado.empty:
            return False, "No se pudo cruzar información de Tools con el Programa. Verifique las entregas.", []

        columnas_finales = ["BOX", "TARA", "BULTOS", "UNI", "LOTE", "SELLO", "RESERVA", "DUS", "MAX"]
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for contrato, data in df_final_agrupado.groupby("Contrato"):
                hoja = str(contrato)
                data_limpia = data[columnas_finales]
                
                data_limpia.to_excel(writer, sheet_name=hoja, index=False, startrow=5)

                ws = writer.sheets[hoja]
                meta = metadata_dict.get(hoja, {})
                
                datos_cabecera = {
                    'nave': meta.get('Nave', ''),
                    'destino': meta.get('DESTINO', ''),
                    'reserva': meta.get('RESERVA', ''),
                    'contrato': hoja,
                    'exportador': "ARAUCO",
                    'embarcador': "CELULOSA ARAUCO",
                    'carga': meta.get('PRODINFO', ''),
                    'linea': meta.get('NAV_CLEAN', '')
                }
                agregar_cabecera_arauco(ws, datos_cabecera)

        wb = load_workbook(output)

        # Merge de celdas BOX para visualización
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            header_row_idx = 6
            idx_box = next((cell.col_idx for cell in ws[header_row_idx] if cell.value == "BOX"), 1)
            
            start = header_row_idx + 1
            max_row = ws.max_row
            
            while start <= max_row:
                valor = ws.cell(row=start, column=idx_box).value
                end = start
                while end + 1 <= max_row and ws.cell(row=end + 1, column=idx_box).value == valor:
                    end += 1
                
                if end > start:
                    ws.merge_cells(start_row=start, start_column=idx_box, end_row=end, end_column=idx_box)
                    ws.cell(row=start, column=idx_box).alignment = Alignment(vertical="center")
                
                start = end + 1

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        return True, "Archivo consolidado generado correctamente", [("RemateCelulosa.xlsx", final_output)]

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []
# ==========================================
#      LÓGICA DE SAG 
# ==========================================
def procesar_sag(rutas):
    st.info("Iniciando procesamiento de SAG...")
    try:
        # 1. Cargar Remate
        path_remate = rutas['remate']
        remate = pd.read_excel(path_remate)
        remate["Contenedor"] = remate["Contenedor"].astype(str).str.strip()
        
        # 2. Cargar Archivos SIF (SAG)
        rutas_sif = rutas['sag']
        if isinstance(rutas_sif, str):
            rutas_sif = [rutas_sif]
            
        lista_sifs = []
        st.info(f"Cargando {len(rutas_sif)} archivos SIF...")
        for ruta in rutas_sif:
            try:
                df_temp = pd.read_excel(ruta, sheet_name="detalle")
                lista_sifs.append(df_temp)
            except Exception as e:
                st.error(f"Error cargando {ruta}: {e}")
        
        if not lista_sifs:
            return False, "No se pudo cargar ningún archivo SIF válido.", []
            
        SAG = pd.concat(lista_sifs, ignore_index=True)
        
        # 3. Cargar Picking
        path_picking = rutas['picking']
        if not os.path.exists(path_picking):
             return False, f"No se encontró el archivo Picking: {path_picking}", []

        picking_pos = pd.read_excel(path_picking, sheet_name="Posicion")
        picking_cab = pd.read_excel(path_picking, sheet_name="Cabecera")

        # 4. Normalizar columnas SAG (SIF)
        if "Codigo_Barra" not in SAG.columns or "SIF" not in SAG.columns:
            return False, "Los archivos SIF no tienen las columnas 'Codigo_Barra' o 'SIF'.", []
            
        SAG["Codigo_Barra"] = SAG["Codigo_Barra"].astype(str).str.strip()
        SAG['SIF_num'] = pd.to_numeric(SAG['SIF'], errors='coerce')
        # Ordenar para quedarnos con el SIF mayor en caso de duplicados
        SAG = SAG.sort_values(by=['Codigo_Barra', 'SIF_num'], ascending=[True, False])
        SAG = SAG.drop_duplicates(subset=['Codigo_Barra'], keep='first')
        SAG['SIF'] = SAG['SIF'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
        
        # 5. Preparar Picking Posición
        picking_pos["Lote"] = picking_pos["Lote"].astype(str).str.strip()
        picking_cab["ID Contenedor"] = picking_cab["ID Contenedor"].astype(str).str.strip()
        
        # Traer el ID Contenedor a la tabla de posiciones (paquetes)
        picking_pos = picking_pos.merge(
            picking_cab[["ID Cabecera", "ID Contenedor"]],
            on="ID Cabecera",
            how="left"
        )
        
        # Cruzar los paquetes con su respectivo SIF
        picking_pos = picking_pos.merge(
            SAG[["Codigo_Barra", "SIF"]],
            how="left",
            left_on="Lote",
            right_on="Codigo_Barra"
        )
        picking_pos["SIF"] = picking_pos["SIF"].fillna("Sin SIF").astype(str).str.strip()

        # =========================================================
        # PROCESO A: Agrupar Picking separando cantidades por SIF
        # =========================================================
        # Primero contamos lotes y sumamos peso por Contenedor + SIF
        resumen_sif = picking_pos.groupby(["ID Contenedor", "SIF"], dropna=False).agg({
            "Lote": "count",
            "Peso": "sum"
        }).reset_index()
        
        # Formateamos para que sean textos limpios
        resumen_sif["SIF"] = resumen_sif["SIF"].astype(str)
        resumen_sif["Lote"] = resumen_sif["Lote"].fillna(0).astype(int).astype(str)
        resumen_sif["Peso"] = resumen_sif["Peso"].fillna(0).round(2).astype(str)
        
        # Luego concatenamos en una sola fila por Contenedor (Ej: "10 / 5")
        sif_agrupado = resumen_sif.groupby("ID Contenedor").agg({
            "SIF": lambda x: " / ".join(x),
            "Lote": lambda x: " / ".join(x),
            "Peso": lambda x: " / ".join(x)
        }).reset_index()
        
        sif_agrupado = sif_agrupado.rename(columns={
            "Lote": "Cantidad de Lotes",
            "Peso": "Peso Lote"
        })

        # =========================================================
        # PROCESO B: Agrupar Remate (1 fila por contenedor)
        # =========================================================
        agg_remate = {
            "Entrega": lambda x: " / ".join(x.dropna().astype(str).unique()),
            "Reserva": "first",
            "Pto Destino": "first",
            "Clase de Producto": "first",
            "Sello": "first",
            "Sello_Inspector": "first",
            "Peso Total (kg)": "sum",
            "Volumen Total (m3)": "sum"
        }
        
        valid_agg = {k: v for k, v in agg_remate.items() if k in remate.columns}
        remate_agrupado = remate.groupby("Contenedor", as_index=False).agg(valid_agg)

        # =========================================================
        # PROCESO C: Unir Remate agrupado con SIF agrupado
        # =========================================================
        remate_final = remate_agrupado.merge(
            sif_agrupado,
            left_on="Contenedor",
            right_on="ID Contenedor",
            how="left"
        )
        
        remate_final = remate_final.drop(columns=["ID Contenedor"], errors="ignore")

        # Ordenar las columnas para el reporte final
        cols_order = [
            "Entrega", "Reserva", "Pto Destino", "Clase de Producto", 
            "Contenedor", "Sello", "Sello Inspector", "SIF", 
            "Cantidad de Lotes", "Peso Lote", "Peso Total (kg)", "Volumen Total (m3)"
        ]
        
        cols_final = [c for c in cols_order if c in remate_final.columns]
        for c in remate_final.columns:
            if c not in cols_final:
                cols_final.append(c)
                
        remate_final = remate_final[cols_final]

        # Creación del Excel en memoria
        output = BytesIO()
        remate_final.to_excel(output, index=False, engine='openpyxl')
        
        wb = load_workbook(output)
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        return True, "Archivo generado correctamente", [("RemateSIF.xlsx", final_output)]

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []
# ==========================================
#      LÓGICA CMPC CELULOSA
# ==========================================
def procesar_cmpc_celulosa(rutas):
    st.info("Iniciando procesamiento CMPC Celulosa...")
    try:
        remate = pd.read_excel(rutas['remate'])
        tools = pd.read_excel(rutas['tools'])

        if 'producto' in remate.columns:
            remate = remate[remate['producto'] != "PAPEL KRAFT"]
        
        if "sello_linea" in remate.columns:
            remate["sello_linea_clean"] = (
                remate["sello_linea"]
                .astype(str)
                .str.replace("-", "", regex=False)
                .str.strip()
            )
        else:
             return False, "Columna 'sello_linea' no encontrada en Remate.", []

        if "Sello_linea" in tools.columns:
            tools["Sello_linea_clean"] = (
                tools["Sello_linea"]
                .astype(str)
                .str.replace("-", "", regex=False)
                .str.strip()
            )
        else:
            return False, "Columna 'Sello_linea' no encontrada en Tools.", []

        sellos_validos = set(remate["sello_linea_clean"])
        tools_filtrado = tools[tools["Sello_linea_clean"].isin(sellos_validos)].copy()

        df = tools_filtrado.merge(
            remate,
            left_on="Sello_linea_clean",
            right_on="sello_linea_clean",
            how="left",
            suffixes=("_tools", "_remate")
        )

        consolidado = pd.DataFrame()
        consolidado["Etiqueta"] = df["Expedicion"]
        consolidado["Contenedor"] = df["Contenedor"]
        consolidado["Sello"] = df["Sello_linea_clean"]
        consolidado["Tara"] = pd.to_numeric(df["Tara"], errors="coerce")
        consolidado["Tipo Cont."] = df["Tipo_Contenedor"]
        consolidado["Dimension"] = df["medida"]
        consolidado["Naviera"] = df["linea"]
        consolidado["Reserva"] = df["reserva"]
        consolidado["Dus"] = df["dus"]
        consolidado["agencia"] = df["aga"]
        consolidado["Bodega"] = ""
        consolidado["ubicación"] = ""
        consolidado["Directo"] = "N"
        consolidado["Destino"] = df["Pto_Destino"].astype(str).str.split(",", n=1).str[0]
        consolidado["Fardos"] = pd.to_numeric(df["Cantidad"], errors="coerce")
        consolidado["Pedido"] = df["Contrato"].astype(str).str.split("-", n=1).str[0]
        consolidado["fecha dus"] = (
            pd.to_datetime(df["fecha_aceptacion"], format="%d/%m/%Y %H:%M", errors="coerce")
            .dt.strftime("%d/%m/%Y")
        )
        consolidado["UNIT"] = consolidado["Fardos"] / 8
        consolidado["PLANTA"] = (
            df["producto"]
            .astype(str)
            .str.upper()
            .str.replace("CELULOSA ", "", regex=False)
            .str.strip()
        )
        consolidado["Peso neto"] = 0.25175 * consolidado["Fardos"]
        consolidado["Peso bruto"] = 0.25413 * consolidado["Fardos"]
        consolidado["Peso Total"] = 24396 + consolidado["Tara"]

        def calcular_volumen(planta, fardos):
            if pd.isna(fardos):
                return np.nan
            planta = str(planta).upper().strip()

            if "STA" in planta or "FÉ" in planta:
                return fardos * 0.254
            elif "LAJA" in planta:
                return fardos * 0.2502
            elif "PACIFICO" in planta:
                return fardos * 0.2618
            return np.nan

        consolidado["Volumen"] = [
            calcular_volumen(p, f)
            for p, f in zip(consolidado["PLANTA"], consolidado["Fardos"])
        ]

        consolidado["Marca"] = consolidado["Etiqueta"].astype(str) + "/" + consolidado["PLANTA"].astype(str)

        output = BytesIO()
        consolidado.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        
        return True, "Archivo generado", [("CMPC_Celulosa_Consolidado.xlsx", output)]

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []

# ==========================================
#      LÓGICA CMPC MADERA (FINAL - NOTA POR CONTENEDOR)
# ==========================================
def procesar_cmpc_madera(rutas):
    st.info("Iniciando procesamiento CMPC Madera...")
    try:
        remate = pd.read_excel(rutas['remate'])
        tools = pd.read_excel(rutas['informe'])

        remate['sigla_cnt'] = remate['sigla_cnt'].astype(str).str.strip()
        remate['nro_cnt'] = remate['nro_cnt'].astype(str).str.strip()
        remate['dv_cnt'] = remate['dv_cnt'].astype(str).str.strip()

        def construir_contenedor_rem(row):
            sigla = str(row['sigla_cnt']).strip()
            val_num = str(row['nro_cnt'])
            if '.' in val_num:
                numero = val_num.split('.')[0].strip()
            else:
                numero = val_num.strip()
            dv = str(row['dv_cnt']).strip()
            numero = numero.zfill(6)
            contenedor = f"{sigla}-{numero}-{dv}"
            return contenedor

        remate['CONTENEDORREM'] = remate.apply(construir_contenedor_rem, axis=1)

        if "Sello_linea" in tools.columns:
            tools["Sello_linea_clean"] = (
                tools["Sello_linea"]
                .astype(str)
                .str.replace("-", "", regex=False)
                .str.strip()
            )

        cols_tools_necesarias = ['Cnt_Sigla', 'Cnt_Nro', 'Cnt_DV']
        for col in cols_tools_necesarias:
            if col not in tools.columns:
                return False, f"El archivo Informe (Tools) no tiene la columna '{col}'", []

        tools['Cnt_Sigla'] = tools['Cnt_Sigla'].astype(str).str.strip()
        tools['Cnt_Nro'] = tools['Cnt_Nro'].astype(str).str.strip()
        tools['Cnt_DV'] = tools['Cnt_DV'].astype(str).str.strip()

        def construir_contenedor_tools(row):
            sigla = str(row['Cnt_Sigla']).strip()
            val_num = str(row['Cnt_Nro'])
            if '.' in val_num:
                numero = val_num.split('.')[0].strip()
            else:
                numero = val_num.strip()
            dv = str(row['Cnt_DV']).strip()
            numero = numero.zfill(6)
            contenedor = f"{sigla}-{numero}-{dv}"
            return contenedor

        tools['CONTENEDORINF'] = tools.apply(construir_contenedor_tools, axis=1)

        mensajes_exito = []
        archivos_output = []
        
        # Limpieza robusta de la columna producto
        if "producto" in remate.columns:
            remate["producto"] = remate["producto"].astype(str).str.strip().str.upper()

        # SUB-PROCESO 1: MADERA SECA
        remate_seca = remate[remate["producto"].astype(str).str.upper().str.contains("SECA", na=False)].copy()
        
        if not remate_seca.empty:
            try:
                remate_seca['Desc_Carga_Calc'] = remate_seca['cant_piezas'].astype(str) + " PIECES, CHILEAN RADIATA PINE"
                
                contenedores_unicos_s = remate_seca['CONTENEDORREM'].unique()
                mapa_nota_s = {cnt: i+1 for i, cnt in enumerate(contenedores_unicos_s)}

                df_remate_extra_seca = pd.DataFrame({
                    "Nota": remate_seca['CONTENEDORREM'].map(mapa_nota_s),
                    "Venta": remate_seca["pedido"],
                    "Reserva": remate_seca["reserva"],
                    "Contenedor": remate_seca["CONTENEDORREM"],
                    "Sello Naviera (Carrier Seal)": remate_seca["sello_linea"],
                    "Descripción de la Carga": remate_seca["Desc_Carga_Calc"],
                    "N° de Pqts.": remate_seca["cant_paquetes"],
                    "Tara del Contenedor": remate_seca["tara"],
                    "Volumen Bruto de la Carga": remate_seca["volumen"],
                    "Peso Bruto de la Carga (documental)": remate_seca["neto"],
                    "Volumen Bruto del Contenedor": remate_seca["volumen"], 
                    "Comentarios del Contenedor": remate_seca["pto_final"]
                })
                
                output_seca_remate = BytesIO()
                df_remate_extra_seca.to_excel(output_seca_remate, index=False, engine='openpyxl')
                output_seca_remate.seek(0)
                archivos_output.append(("Remate_CMPC_Madera_Seca.xlsx", output_seca_remate))
                
            except Exception as e:
                st.warning(f"Error generando Remate Extra Seca: {e}")

            # CONSOLIDADO SECA
            tools_filtrado = tools[tools['CONTENEDORINF'].isin(remate_seca['CONTENEDORREM'])]
            remate_matched = remate_seca.set_index("CONTENEDORREM")
            tools_matched = tools_filtrado.set_index("CONTENEDORINF")
            
            df = tools_matched.join(remate_matched, how="left", rsuffix="_rem")
            
            if not df.empty:
                split_pedido = df['Orden_Pedido'].astype(str).str.split('-', n=1)
                df['contrato'] = split_pedido.str[0]
                df['item'] = split_pedido.str[1]
                df['fecha_dus'] = pd.to_datetime(df['fecha_aceptacion'], errors='coerce').dt.strftime('%d/%m/%Y')

                df_consolidado = pd.DataFrame({
                    "Npaquete": df["Nro_Paquete"],
                    "Contenedor": df.index,
                    "Sello": df["sello_linea"],
                    "Tara": df["tara"],
                    "Dimension": df["medida"],
                    "Tipo Cont.": df["tipo"],
                    "Directo": "N",
                    "Destino": df["pto_final"],
                    "Paquetes": "1",
                    "contrato": df["contrato"],
                    "item": df["item"],
                    "Naviera": df["linea"],
                    "Bodega": "",
                    "ubicación": "",
                    "Reserva": df["reserva"],
                    "Dus": df["dus"],
                    "fecha dus": df["fecha_dus"],
                    "agencia": df["aga"],
                })

                output_seca_cons = BytesIO()
                df_consolidado.to_excel(output_seca_cons, index=False, engine='openpyxl')
                output_seca_cons.seek(0)
                archivos_output.append(("CMPC_Madera_Seca_Consolidado.xlsx", output_seca_cons))
                # =========================================================
                # 2. NUEVO: CONSOLIDADO SAG (Nivel paquete + columnas extra)
                # =========================================================
                df_consolidado_sag = df_consolidado.copy()
                
                # Extraemos los datos provenientes de la tabla "df" (Tools cruzado con Remate).
                # Nota: Asegúrate de que los nombres "Peso_lote", "Volumen_Lote" y "Sello_Inspector" 
                # coincidan exactamente con cómo vienen en la cabecera del Excel "Tools" de CMPC.
                
                df_consolidado_sag["Peso"] = pd.to_numeric(df.get("Peso_lote", df.get("peso", 0)), errors="coerce").fillna(0)
                df_consolidado_sag["Volumen"] = pd.to_numeric(df.get("Volumen_Lote", df.get("volumen_tools", 0)), errors="coerce").fillna(0)
                df_consolidado_sag["Sello Inspector"] = df.get("Sello_Inspector", "")

                # Exportamos a un nuevo archivo Excel
                output_seca_sag = BytesIO()
                df_consolidado_sag.to_excel(output_seca_sag, index=False, engine='openpyxl')
                output_seca_sag.seek(0)
                
                # Lo agregamos al arreglo final para que se descargue
                archivos_output.append(("CMPC_Madera_Seca_Consolidado_SAG.xlsx", output_seca_sag))
                
        # SUB-PROCESO 2: MADERA VERDE
        remate_verde = remate[remate["producto"].astype(str).str.upper().str.contains("VERDE", na=False)].copy()
        
        if not remate_verde.empty:
            try:
                remate_verde['Desc_Carga_Calc'] = remate_verde['cant_piezas'].astype(str) + " PIECES, CHILEAN RADIATA PINE"
                
                contenedores_unicos_v = remate_verde['CONTENEDORREM'].unique()
                mapa_nota_v = {cnt: i+1 for i, cnt in enumerate(contenedores_unicos_v)}

                df_remate_extra_verde = pd.DataFrame({
                    "Nota": remate_verde['CONTENEDORREM'].map(mapa_nota_v),
                    "Venta": remate_verde["pedido"],
                    "Reserva": remate_verde["reserva"],
                    "Contenedor": remate_verde["CONTENEDORREM"],
                    "Sello Naviera (Carrier Seal)": remate_verde["sello_linea"],
                    "Descripción de la Carga": remate_verde["Desc_Carga_Calc"],
                    "N° de Pqts.": remate_verde["cant_paquetes"],
                    "Tara del Contenedor": remate_verde["tara"],
                    "Volumen Bruto de la Carga": remate_verde["volumen"],
                    "Peso Bruto de la Carga (documental)": remate_verde["neto"],
                    "Volumen Bruto del Contenedor": remate_verde["volumen"],
                    "Comentarios del Contenedor": remate_verde["pto_final"]
                })
                
                output_verde_remate = BytesIO()
                df_remate_extra_verde.to_excel(output_verde_remate, index=False, engine='openpyxl')
                output_verde_remate.seek(0)
                archivos_output.append(("Remate_CMPC_Madera_Verde.xlsx", output_verde_remate))
                
            except Exception as e:
                st.warning(f"Error generando Remate Extra Verde: {e}")

            # CONSOLIDADO VERDE
            tools_filtrado_v = tools[tools['CONTENEDORINF'].isin(remate_verde['CONTENEDORREM'])]
            remate_matched_v = remate_verde.set_index("CONTENEDORREM")
            tools_matched_v = tools_filtrado_v.set_index("CONTENEDORINF")
            
            df_v = tools_matched_v.join(remate_matched_v, how="left", rsuffix="_rem")
            
            if not df_v.empty:
                split_pedido_v = df_v['Orden_Pedido'].astype(str).str.split('-', n=1)
                df_v['contrato'] = split_pedido_v.str[0]
                df_v['item'] = split_pedido_v.str[1]
                df_v['fecha_dus'] = pd.to_datetime(df_v['fecha_aceptacion'], errors='coerce').dt.strftime('%d/%m/%Y')

                df_consolidado_v = pd.DataFrame({
                    "Npaquete": df_v["Nro_Paquete"],
                    "Contenedor": df_v.index,
                    "Sello": df_v["sello_linea"],
                    "Tara": df_v["tara"],
                    "Dimension": df_v["medida"],
                    "Tipo Cont.": df_v["tipo"],
                    "Directo": "N",
                    "Destino": df_v["pto_final"],
                    "Paquetes": "1",
                    "contrato": df_v["contrato"],
                    "item": df_v["item"],
                    "Naviera": df_v["linea"],
                    "Bodega": "",
                    "ubicación": "",
                    "Reserva": df_v["reserva"],
                    "Dus": df_v["dus"],
                    "fecha dus": df_v["fecha_dus"],
                    "agencia": df_v["aga"],
                })

                output_verde_cons = BytesIO()
                df_consolidado_v.to_excel(output_verde_cons, index=False, engine='openpyxl')
                output_verde_cons.seek(0)
                archivos_output.append(("CMPC_Madera_Verde_Consolidado.xlsx", output_verde_cons))
                # =========================================================
                # 2. NUEVO: CONSOLIDADO SAG (Nivel paquete + columnas extra)
                # =========================================================
                df_consolidado_sag = df_consolidado.copy()
                
                # Extraemos los datos provenientes de la tabla "df" (Tools cruzado con Remate).
                # Nota: Asegúrate de que los nombres "Peso_lote", "Volumen_Lote" y "Sello_Inspector" 
                # coincidan exactamente con cómo vienen en la cabecera del Excel "Tools" de CMPC.
                
                df_consolidado_sag["Peso"] = pd.to_numeric(df.get("Peso_lote", df.get("peso", 0)), errors="coerce").fillna(0)
                df_consolidado_sag["Volumen"] = pd.to_numeric(df.get("Volumen_Lote", df.get("volumen_tools", 0)), errors="coerce").fillna(0)
                df_consolidado_sag["Sello Inspector"] = df.get("Sello_Inspector", "")

                # Exportamos a un nuevo archivo Excel
                output_seca_sag = BytesIO()
                df_consolidado_sag.to_excel(output_seca_sag, index=False, engine='openpyxl')
                output_seca_sag.seek(0)
                
                # Lo agregamos al arreglo final para que se descargue
                archivos_output.append(("CMPC_Madera_Seca_Consolidado_SAG.xlsx", output_seca_sag))
        if not archivos_output:
            return True, "Proceso finalizado, pero no se generaron archivos.", []

        return True, "Archivos generados exitosamente", archivos_output

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []

# ==========================================
#      LÓGICA CMPC PAPEL (FINAL - NOTA POR CONTENEDOR)
# ==========================================
def procesar_cmpc_papel(rutas):
    st.info("Iniciando procesamiento CMPC Papel...")
    try:
        remate = pd.read_excel(rutas['remate'])
        tools = pd.read_excel(rutas['tools'])

        archivos_output = []

        # 1. NORMALIZACIÓN DE COLUMNAS Y CONTENEDORES
        remate.columns = [c.strip() for c in remate.columns]
        
        col_tara_rem = next((c for c in remate.columns if c.lower() == 'tara'), 'tara')
        col_pto_rem = next((c for c in remate.columns if c.lower() in ['pto_descarga', 'pto_final', 'puerto_destino']), 'pto_descarga')
        
        remate['sigla_cnt'] = remate['sigla_cnt'].astype(str).str.strip()
        remate['nro_cnt'] = remate['nro_cnt'].astype(str).str.strip()
        remate['dv_cnt'] = remate['dv_cnt'].astype(str).str.strip()

        def construir_contenedor_rem(row):
            sigla = str(row['sigla_cnt']).strip()
            val_num = str(row['nro_cnt']).split('.')[0].strip()
            dv = str(row['dv_cnt']).strip()
            return f"{sigla}-{val_num.zfill(6)}-{dv}"

        remate['CONTENEDORREM'] = remate.apply(construir_contenedor_rem, axis=1)

        tools.columns = [c.strip() for c in tools.columns]
        
        tools['Cnt_Sigla'] = tools['Cnt_Sigla'].astype(str).str.strip()
        tools['Cnt_Nro'] = tools['Cnt_Nro'].astype(str).str.strip()
        tools['Cnt_DV'] = tools['Cnt_DV'].astype(str).str.strip()
        
        col_sello_tools = next((c for c in tools.columns if c.lower() == 'sello_linea'), 'Sello_linea')
        if col_sello_tools in tools.columns:
            tools["Sello_linea_clean"] = tools[col_sello_tools].astype(str).str.replace("-", "", regex=False).str.strip()
        else:
            tools["Sello_linea_clean"] = ""

        col_peso_tools = next((c for c in tools.columns if c.lower() == 'peso_lote'), None)
        if col_peso_tools:
            tools[col_peso_tools] = tools[col_peso_tools].astype(str).str.replace(',', '.', regex=False)
            tools[col_peso_tools] = pd.to_numeric(tools[col_peso_tools], errors='coerce').fillna(0)
        else:
            tools['Peso_lote'] = 0
            col_peso_tools = 'Peso_lote'

        def construir_contenedor_tools(row):
            sigla = str(row['Cnt_Sigla']).strip()
            val_num = str(row['Cnt_Nro']).split('.')[0].strip()
            dv = str(row['Cnt_DV']).strip()
            return f"{sigla}-{val_num.zfill(6)}-{dv}"

        tools['CONTENEDORINF'] = tools.apply(construir_contenedor_tools, axis=1)

        # 2. GENERAR ARCHIVO NUEVO "REMATE_CMPC_PAPEL"
        try:
            grupo_tools = tools.groupby(['Orden_Pedido', 'CONTENEDORINF']).agg({
                'Reserva': 'first',
                'Sello_linea_clean': 'first',
                'Nro_Paquete': 'count',
                col_peso_tools: 'sum'
            }).reset_index()

            remate_subset = remate[['CONTENEDORREM', col_tara_rem, col_pto_rem]].drop_duplicates('CONTENEDORREM')
            
            df_nuevo = grupo_tools.merge(
                remate_subset,
                left_on='CONTENEDORINF',
                right_on='CONTENEDORREM',
                how='left'
            )

            contenedores_unicos = df_nuevo['CONTENEDORINF'].unique()
            mapa_id_contenedor = {cnt: i+1 for i, cnt in enumerate(contenedores_unicos)}
            
            df_exportar = pd.DataFrame()
            df_exportar['Nota'] = df_nuevo['CONTENEDORINF'].map(mapa_id_contenedor)
            df_exportar['Número Venta'] = df_nuevo['Orden_Pedido']
            df_exportar['Reserva'] = df_nuevo['Reserva']
            df_exportar['Contenedor'] = df_nuevo['CONTENEDORINF']
            df_exportar['Sello Naviera (Carrier Seal)'] = df_nuevo['Sello_linea_clean']
            df_exportar['Descripción de la Carga'] = "PAPEL KRAFT"
            df_exportar['N° de Pqts.'] = df_nuevo['Nro_Paquete']
            df_exportar['Tara del Contenedor'] = df_nuevo[col_tara_rem]
            df_exportar['Peso Bruto de la Carga (documental)'] = df_nuevo[col_peso_tools]
            df_exportar['Comentarios del Contenedor'] = df_nuevo[col_pto_rem]

            output_remate = BytesIO()
            df_exportar.to_excel(output_remate, index=False, engine='openpyxl')
            output_remate.seek(0)
            archivos_output.append(("Remate_CMPC_Papel.xlsx", output_remate))

        except Exception as e:
            st.warning(f"Error generando Remate Nuevo: {e}")

        # 3. GENERAR ARCHIVO ANTIGUO "CONSOLIDADO"
        try:
            remate_papel = remate[remate["producto"] == "PAPEL KRAFT"].copy()
            tools_filt = tools[tools['CONTENEDORINF'].isin(remate_papel['CONTENEDORREM'])].copy()
            
            df_cons = tools_filt.set_index("CONTENEDORINF").join(
                remate_papel.set_index("CONTENEDORREM"), 
                how="left", 
                rsuffix="_rem"
            )

            if not df_cons.empty:
                df_cons['fecha_dus'] = pd.to_datetime(df_cons['fecha_aceptacion'], errors='coerce').dt.strftime('%d/%m/%Y')
                
                df_consolidado_final = pd.DataFrame({
                    "Etiqueta": df_cons["Nro_Paquete"], 
                    "Contenedor": df_cons.index, 
                    "Sello": df_cons["sello_linea"],
                    "Tara": df_cons[col_tara_rem], 
                    "Dimension": df_cons["medida"], 
                    "Tipo Cont.": df_cons["tipo"], 
                    "Directo": "N",
                    "Destino": df_cons["pto_final"], 
                    "Fardos": "1", 
                    "contrato": df_cons["Orden_Pedido"], 
                    "item": "10",
                    "Naviera": df_cons["linea"], 
                    "Bodega": "", 
                    "ubicación": "", 
                    "Reserva": df_cons["reserva"], 
                    "Dus": df_cons["dus"], 
                    "fecha dus": df_cons["fecha_dus"], 
                    "agencia": df_cons["aga"]
                })

                output_consolidado = BytesIO()
                df_consolidado_final.to_excel(output_consolidado, index=False, engine='openpyxl')
                output_consolidado.seek(0)
                archivos_output.append(("CMPC_Papel_Consolidado.xlsx", output_consolidado))

        except Exception as e:
            st.warning(f"Error generando Consolidado: {e}")

        if not archivos_output:
            return True, "Proceso finalizado, pero no se generaron archivos.", []

        return True, "Archivos generados exitosamente", archivos_output

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []

# ==========================================
#      LÓGICA CMPC PLYWOOD (FINAL - NOTA POR CONTENEDOR)
# ==========================================
def procesar_cmpc_plywood(rutas):
    st.info("Iniciando procesamiento CMPC Plywood...")
    try:
        remate = pd.read_excel(rutas['remate'])
        tools = pd.read_excel(rutas['tools'])

        remate['sigla_cnt'] = remate['sigla_cnt'].astype(str).str.strip()
        remate['nro_cnt'] = remate['nro_cnt'].astype(str).str.strip()
        remate['dv_cnt'] = remate['dv_cnt'].astype(str).str.strip()
        
        tools['Cnt_Sigla'] = tools['Cnt_Sigla'].astype(str).str.strip()
        tools['Cnt_Nro'] = tools['Cnt_Nro'].astype(str).str.strip()
        tools['Cnt_DV'] = tools['Cnt_DV'].astype(str).str.strip()
        
        if "Sello_linea" in tools.columns:
            tools["Sello_linea_clean"] = tools["Sello_linea"].astype(str).str.replace("-", "", regex=False).str.strip()

        def construir_contenedor(row):
            sigla = str(row['sigla_cnt']).strip()
            val_num = str(row['nro_cnt'])
            if '.' in val_num: numero = val_num.split('.')[0].strip()
            else: numero = val_num.strip()
            dv = str(row['dv_cnt']).strip()
            numero = numero.zfill(6)
            contenedor = f"{sigla}-{numero}-{dv}"
            return contenedor

        def construir_contenedor2(row):
            sigla = str(row['Cnt_Sigla']).strip()
            val_num = str(row['Cnt_Nro'])
            if '.' in val_num: numero = val_num.split('.')[0].strip()
            else: numero = val_num.strip()
            dv = str(row['Cnt_DV']).strip()
            numero = numero.zfill(6)
            contenedor = f"{sigla}-{numero}-{dv}"
            return contenedor

        remate['CONTENEDORREM'] = remate.apply(construir_contenedor, axis=1)
        tools['CONTENEDORINF'] = tools.apply(construir_contenedor2, axis=1)

        archivos_output = []

        remate_ply = remate[remate["producto"] == "PLYWOOD"].copy()
        
        if remate_ply.empty: 
            return True, "No se encontraron registros con producto 'PLYWOOD' en el archivo Remate.", []

        # GENERAR REMATE EXTRA
        try:
            remate_ply['Desc_Carga_Calc'] = remate_ply['cant_piezas'].astype(str) + " PIECES, PLYWOOD"
            
            contenedores_unicos = remate_ply['CONTENEDORREM'].unique()
            mapa_nota = {cnt: i+1 for i, cnt in enumerate(contenedores_unicos)}
            
            df_remate_extra = pd.DataFrame({
                "Nota": remate_ply['CONTENEDORREM'].map(mapa_nota),
                "Venta": remate_ply["pedido"],
                "Reserva": remate_ply["reserva"],
                "Contenedor": remate_ply["CONTENEDORREM"],
                "Sello Naviera (Carrier Seal)": remate_ply["sello_linea"],
                "Descripción de la Carga": remate_ply['Desc_Carga_Calc'],
                "N° de Pqts.": remate_ply["cant_paquetes"],
                "Tara del Contenedor": remate_ply["tara"],
                "Volumen Bruto de la Carga": remate_ply["volumen"],
                "Peso Bruto de la Carga (documental)": remate_ply["neto"],
                "Volumen Bruto del Contenedor": remate_ply["volumen"],
                "Comentarios del Contenedor": remate_ply["pto_final"]
            })
            
            output_remate = BytesIO()
            df_remate_extra.to_excel(output_remate, index=False, engine='openpyxl')
            output_remate.seek(0)
            archivos_output.append(("Remate_CMPC_Plywood.xlsx", output_remate))
            
        except Exception as e:
            st.warning(f"Error generando Remate Extra Plywood: {e}")

        # LÓGICA ORIGINAL: CONSOLIDADO
        tools_filt = tools[tools['CONTENEDORINF'].isin(remate_ply['CONTENEDORREM'])]
        
        df = tools_filt.set_index("CONTENEDORINF").join(remate_ply.set_index("CONTENEDORREM"), how="left", rsuffix="_rem")
        
        if not df.empty:
            df['fecha_dus'] = pd.to_datetime(df['fecha_aceptacion'], errors='coerce').dt.strftime('%d/%m/%Y')
            
            df_consolidado = pd.DataFrame({
                "Npaquete": df["Nro_Paquete"], 
                "Contenedor": df.index, 
                "Sello": df["sello_linea"],
                "Tara": df["tara"], 
                "Dimension": df["medida"], 
                "Tipo Cont.": df["tipo"], 
                "Directo": "N",
                "Destino": df["pto_final"], 
                "Fardos": "1", 
                "contrato": df["Orden_Pedido"], 
                "item": "10",
                "Naviera": df["linea"], 
                "Bodega": "", 
                "ubicación": "", 
                "Reserva": df["reserva"], 
                "Dus": df["dus"], 
                "fecha dus": df["fecha_dus"], 
                "agencia": df["aga"]
            })

            output_consolidado = BytesIO()
            df_consolidado.to_excel(output_consolidado, index=False, engine='openpyxl')
            output_consolidado.seek(0)
            archivos_output.append(("CMPC_Plywood_Consolidado.xlsx", output_consolidado))

        if not archivos_output:
            return True, "Proceso finalizado sin generar archivos.", []

        return True, "Archivos generados exitosamente", archivos_output

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []
# ==========================================
#      LÓGICA GENERAL: CUADRATURA CELULOSA
# ==========================================
def procesar_cuadratura_celulosa(rutas):
    st.info("Iniciando Cuadratura de Celulosa...")
    try:
        rutas_bodegas = rutas['bodegas']
        rutas_sistema = rutas['sistema']
        
        # Asegurar que sean listas
        if isinstance(rutas_bodegas, str): rutas_bodegas = [rutas_bodegas]
        if isinstance(rutas_sistema, str): rutas_sistema = [rutas_sistema]
        
        # ==========================================
        # 1. PROCESAR DATOS FÍSICOS (PLANOS DE BODEGA)
        # ==========================================
        datos_fisico = []
        for ruta in rutas_bodegas:
            nombre_original = os.path.basename(ruta).split('_')[-1]
            bodega_nombre = nombre_original.split('.')[0].upper()
            
            try:
                dict_hojas = pd.read_excel(ruta, header=None, sheet_name=None)
                for nombre_hoja, df in dict_hojas.items():
                    rows, cols = df.shape
                    puntos_inicio = []
                    
                    for r in range(rows):
                        for c in range(cols - 4):
                            try:
                                h1 = str(df.iloc[r, c]).strip().upper()
                                h2 = str(df.iloc[r, c+1]).strip().upper()
                                if 'DESCRIPCI' in h1 and 'LOTE' in h2:
                                    puntos_inicio.append((r, c))
                            except: continue
                    
                    for r_start, c_start in puntos_inicio:
                        r_actual = r_start + 1
                        consecutivos_vacios = 0
                        
                        while r_actual < rows:
                            fila_data = df.iloc[r_actual, c_start:c_start+5]
                            vals_fila = [str(x) for x in fila_data.values if pd.notna(x)]
                            
                            if len("".join(vals_fila).strip()) < 2:
                                consecutivos_vacios += 1
                                if consecutivos_vacios >= 10: break
                                r_actual += 1
                                continue
                            
                            consecutivos_vacios = 0
                            desc = str(fila_data.iloc[0]).strip()
                            lote = str(fila_data.iloc[1]).strip()
                            
                            if lote.endswith('.0'): lote = lote[:-2]
                            lote = re.sub(r'\D', '', lote)
                            
                            units_raw = fila_data.iloc[3]
                            plano_estado = str(fila_data.iloc[4]).strip().upper()
                            
                            if desc == 'nan' or 'DESCRIPCI' in desc.upper():
                                r_actual += 1
                                continue
                                
                            estado_final = "DAÑADA" if plano_estado == 'DAÑADA' else "NORMAL"
                            
                            try:
                                val_str = str(units_raw).replace(',', '.')
                                units = float(val_str)
                                fardos = units * 8
                            except:
                                units, fardos = 0, 0
                                
                            if pd.notna(fila_data.iloc[1]) and lote != 'nan' and lote != '':
                                datos_fisico.append({
                                    'Bodega': bodega_nombre,
                                    'Lote': lote,
                                    'Cliente/Desc': desc,
                                    'Unit': units,
                                    'Fardos': fardos,
                                    'Estado': estado_final
                                })
                            r_actual += 1
            except Exception as e:
                st.warning(f"Error procesando plano {nombre_original}: {e}")

        df_fisico = pd.DataFrame(datos_fisico)

         # ==========================================
        # 2. PROCESAR DATOS DEL SISTEMA (Múltiples Archivos)
        # ==========================================
        lista_df_sistemas = []
        
        for ruta_sys in rutas_sistema:
            df_temp = pd.DataFrame()
            
            # Apertura robusta multi-formato
            try:
                df_temp = pd.read_excel(ruta_sys)
            except Exception:
                try:
                    df_temp = pd.read_html(ruta_sys, decimal=',', thousands='.')[0]
                except Exception:
                    try:
                        df_temp = pd.read_csv(ruta_sys, sep='\t', encoding='latin-1')
                    except Exception:
                        try:
                            df_temp = pd.read_csv(ruta_sys, sep='\t', encoding='utf-16-le')
                        except Exception as e:
                            st.warning(f"No se pudo abrir uno de los archivos: {e}")
            
            if not df_temp.empty:
                df_temp.columns = df_temp.columns.astype(str).str.strip().str.upper()
                lista_df_sistemas.append(df_temp)

        # --- Consolidación de Datos del Sistema ---
        try:
            if lista_df_sistemas:
                df_sistema_raw = pd.concat(lista_df_sistemas, ignore_index=True)
                df_sistema = pd.DataFrame()
                
                # Búsqueda de columna Expedición
                col_exp = next((c for c in df_sistema_raw.columns if 'EXPEDICI' in c or 'LOTE' in c), None)
                if not col_exp:
                    raise ValueError("No se encontró la columna 'Expedición' en los archivos base.")

                # ==========================================
                # FIX: Operaciones vectorizadas seguras (Adiós error de float)
                # ==========================================
                df_sistema['Lote'] = df_sistema_raw[col_exp].astype(str).str.strip()
                df_sistema['Lote'] = df_sistema['Lote'].str.replace(r'\.0$', '', regex=True) # Quita ".0" final si existe
                df_sistema['Lote'] = df_sistema['Lote'].str.replace(r'\D', '', regex=True)   # Mantiene solo números
                
                # Filtrar vacíos y celdas NaN que hayan quedado como texto
                df_sistema = df_sistema[(df_sistema['Lote'] != '') & (df_sistema['Lote'].str.lower() != 'nan')]
                
                # Buscar Bodega
                col_bod = next((c for c in df_sistema_raw.columns if 'BODEGA' in c), None)
                df_sistema['Bodega'] = df_sistema_raw[col_bod].astype(str) if col_bod else 'N/A'
                
                # Buscar Fecha Máxima
                col_fmax = next((c for c in df_sistema_raw.columns if 'MÁX' in c or 'MAX' in c or 'RECEPCI' in c), None)
                df_sistema['F. Máx. Recepción'] = df_sistema_raw[col_fmax] if col_fmax else 'N/A'
                
                # Buscar Cantidad/Stock principal (Sano)
                col_stock = next((c for c in df_sistema_raw.columns if 'STOCK' in c and 'BLOQ' not in c and 'DAÑA' not in c), None)
                if not col_stock: # Fallback por si la columna se llama solo "STOCK"
                    col_stock = next((c for c in df_sistema_raw.columns if 'STOCK' in c), None)

                if col_stock:
                    valores_stock = df_sistema_raw[col_stock].astype(str).str.replace(',', '.', regex=False)
                    df_sistema['Fardos'] = pd.to_numeric(valores_stock, errors='coerce').fillna(0)
                else:
                    df_sistema['Fardos'] = 0
                    
                df_sistema['Unit'] = df_sistema['Fardos'] / 8

                # LÓGICA PARA STOCK DAÑADO EN SISTEMA
                # 1. Puede venir en una columna de cantidad aparte (ej: "Stock Bloqueado" o "Stock Dañado")
                col_stock_bloq = next((c for c in df_sistema_raw.columns if 'BLOQ' in c or 'DAÑA' in c or 'RECHAZ' in c), None)
                if col_stock_bloq:
                    valores_bloq = df_sistema_raw[col_stock_bloq].astype(str).str.replace(',', '.', regex=False)
                    df_sistema['Unit Dañado'] = pd.to_numeric(valores_bloq, errors='coerce').fillna(0) / 8
                else:
                    # 2. O puede venir especificado en una columna de "Estado"
                    col_estado = next((c for c in df_sistema_raw.columns if 'ESTADO' in c or 'CONDICI' in c), None)
                    if col_estado:
                        estado_sys = df_sistema_raw[col_estado].astype(str).str.upper()
                        mask_danado = estado_sys.str.contains('DAÑA|BLOQ|RECH|MAL', na=False)
                        
                        df_sistema['Unit Dañado'] = np.where(mask_danado, df_sistema['Unit'], 0)
                        # Dejamos solo lo sano en Unit normal
                        df_sistema['Unit'] = np.where(~mask_danado, df_sistema['Unit'], 0) 
                    else:
                        df_sistema['Unit Dañado'] = 0
                
            else:
                df_sistema = pd.DataFrame(columns=['Lote', 'Bodega', 'F. Máx. Recepción', 'Fardos', 'Unit', 'Unit Dañado'])
                
        except Exception as e:
            st.error(f"Error al armar la tabla del sistema: {e}")
            df_sistema = pd.DataFrame(columns=['Lote', 'Bodega', 'F. Máx. Recepción', 'Fardos', 'Unit', 'Unit Dañado'])

        # ==========================================
        # 3. GENERAR CRUCE (Agrupación y Cuadratura)
        # ==========================================
        if not df_fisico.empty:
            # Separar Sano y Dañado en el físico mediante columnas
            df_fisico['Unit Sano'] = np.where(df_fisico['Estado'] != 'DAÑADA', df_fisico['Unit'], 0)
            df_fisico['Unit Dañado'] = np.where(df_fisico['Estado'] == 'DAÑADA', df_fisico['Unit'], 0)
            
            df_fisico_agrupado = df_fisico.groupby('Lote', as_index=False).agg({
                'Cliente/Desc': 'first', 
                'Unit Sano': 'sum',
                'Unit Dañado': 'sum'
            }).rename(columns={'Unit Sano': 'Unit Físico', 'Unit Dañado': 'Stock Físico Dañado'})
        else:
            df_fisico_agrupado = pd.DataFrame(columns=['Lote', 'Cliente/Desc', 'Unit Físico', 'Stock Físico Dañado'])

        if not df_sistema.empty:
            # Sumar Sano y Dañado en el sistema
            df_sistema_agrupado = df_sistema.groupby('Lote', as_index=False).agg({
                'F. Máx. Recepción': 'first', 
                'Unit': 'sum',
                'Unit Dañado': 'sum'
            }).rename(columns={'Unit': 'Unit Sistema', 'Unit Dañado': 'Stock Dañado Sistema'})
        else:
            df_sistema_agrupado = pd.DataFrame(columns=['Lote', 'F. Máx. Recepción', 'Unit Sistema', 'Stock Dañado Sistema'])

        df_cruce = pd.merge(df_fisico_agrupado, df_sistema_agrupado, on='Lote', how='outer')
        
        # Rellenar nulos con 0 para realizar operaciones matemáticas sin error
        cols_numericas = ['Unit Físico', 'Stock Físico Dañado', 'Unit Sistema', 'Stock Dañado Sistema']
        for col in cols_numericas:
            df_cruce[col] = df_cruce[col].fillna(0)
            
        df_cruce['Cliente/Desc'] = df_cruce['Cliente/Desc'].fillna('Solo en Sistema')
        df_cruce['F. Máx. Recepción'] = df_cruce['F. Máx. Recepción'].fillna('Solo en Físico')
        
        # La cuadratura/diferencia principal sigue comparando solo el stock SANO
        df_cruce['Diferencia'] = df_cruce['Unit Físico'] - df_cruce['Unit Sistema']
        df_cruce['Cuadrado'] = df_cruce['Diferencia'].apply(lambda x: 'Si' if abs(x) < 0.01 else 'No')
        
        # Agregamos las nuevas columnas de daño al final del reporte
        columnas_ordenadas = [
            'Lote', 'Cliente/Desc', 'F. Máx. Recepción', 
            'Unit Físico', 'Unit Sistema', 'Diferencia', 'Cuadrado',
            'Stock Físico Dañado', 'Stock Dañado Sistema'
        ]
        df_cruce = df_cruce[columnas_ordenadas].sort_values(by='Cuadrado', ascending=False)

        # ==========================================
        # 4. EXPORTAR A EXCEL EN MEMORIA
        # ==========================================
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if not df_fisico.empty:
                df_fisico.to_excel(writer, sheet_name='Físico', index=False)
            else:
                pd.DataFrame(['Sin datos']).to_excel(writer, sheet_name='Físico', index=False)
                
            if not df_sistema.empty:
                df_sistema.to_excel(writer, sheet_name='Sistema', index=False)
            else:
                pd.DataFrame(['Sin datos']).to_excel(writer, sheet_name='Sistema', index=False)
                
            if not df_cruce.empty:
                df_cruce.to_excel(writer, sheet_name='Cruce', index=False)
            else:
                pd.DataFrame(['Sin datos']).to_excel(writer, sheet_name='Cruce', index=False)
                
        output.seek(0)
        return True, "Cuadratura generada exitosamente", [("Cuadratura_Celulosa.xlsx", output)]

    except Exception as e:
        st.error(f"Error en procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, str(e), []
# ==========================================
#      INTERFAZ STREAMLIT
# ==========================================
CONFIG_ARCHIVOS = {
    "Madera": [
        {"id": "programa", "nombre": "Programa", "opcional": False, "descripcion": "Último programa de consolidación"},
        {"id": "saldos",   "nombre": "Saldos",   "opcional": True},
        {"id": "historico","nombre": "Remates Ant.", "opcional": True, "multiple": True},
        {"id": "tools",    "nombre": "Tools",    "opcional": False, "descripcion": "Subir tools con CB (Contrato,Contenedor,Expedicion,Tara,Cantidad,Sello_linea,Reserva,Orden_Embarque,Orden_Pedido,Cnt_Sigla,Cnt_DV,Cnt_Nro,Marca"},
        {"id": "zoopp",    "nombre": "Zoopp",    "opcional": False, "descripcion": "Consulta ZOOPP SAP"},
    ],
    "Celulosa": [
        {"id": "programa", "nombre": "Programa", "opcional": False, "descripcion": "Programa que contiene entregas BKP/EKP/UKP y/o DP"},
        {"id": "saldos",   "nombre": "Saldos",   "opcional": True},
        {"id": "tools",    "nombre": "Tools",    "opcional": False, "multiple": True, "descripcion": "Subir tools con/sin CB (Contrato,Contenedor,Expedicion,Tara,Cantidad,Sello_linea,Reserva,Orden_Embarque,Orden_Pedido,Cnt_Sigla,Cnt_DV,Cnt_Nro,Marca"},
        {"id": "historico","nombre": "Remates Ant.", "opcional": True, "multiple": True},
    ],
    "SAG": [
        {"id": "remate", "nombre": "Remate", "opcional": False},
        {"id": "picking", "nombre": "Picking", "opcional": False},
        {"id": "sag",   "nombre": "SIF",   "opcional": False, "multiple": True}
    ],
    "CMPC Celulosa": [
        {"id": "remate", "nombre": "Remate", "opcional": False},
        {"id": "tools",  "nombre": "Tools",  "opcional": False},
    ],
    "CMPC Madera": [
        {"id": "remate", "nombre": "Remate", "opcional": False, "descripcion": "Prog. Consolidación -> Consolidaciones con Transmiciones Electrónicas"},
        {"id": "informe", "nombre": "Tools", "opcional": False, "descripcion": "Tools -> Despacho a Contenedor con Código de Barra"},
    ],
    "CMPC Papel": [
        {"id": "remate", "nombre": "Remate", "opcional": False},
        {"id": "tools",  "nombre": "Tools",  "opcional": False},
    ],
    "CMPC Plywood": [
        {"id": "remate", "nombre": "Remate", "opcional": False},
        {"id": "tools",  "nombre": "Tools",  "opcional": False},
    ],  
    "Cuadrar Celulosa": [
        {"id": "bodegas", "nombre": "Planos de Bodega", "opcional": False, "multiple": True, "descripcion": "Archivos Excel con el plano físico de las bodegas."},
        {"id": "sistema", "nombre": "Stock Sistema", "opcional": False, "multiple": True, "descripcion": "Consultas -> Consulta Stock sin Código de Barra"}
    ]    
}

def get_file_uploader_key(file_id, session_id):
    return f"{file_id}_{session_id}"

def aplicar_estilos():
    st.markdown("""
        <style>
        /* Animaciones y estilo para los botones normales */
        div.stButton > button:first-child {
            border-radius: 10px;
            font-weight: 600;
            transition: all 0.3s ease-in-out;
            border: 1px solid #d1d5db;
        }
        div.stButton > button:first-child:hover {
            border-color: #3b82f6;
            color: #3b82f6;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
            transform: translateY(-2px);
        }
        
        /* Estilo especial para el botón de "🚀 Ejecutar Proceso" (Form Submit) */
        div.stFormSubmitButton > button:first-child {
            background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
            color: white;
            border: none;
            border-radius: 10px;
            font-weight: bold;
            transition: all 0.3s ease;
        }
        div.stFormSubmitButton > button:first-child:hover {
            background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%);
            box-shadow: 0 10px 15px -3px rgba(59, 130, 246, 0.4);
            transform: translateY(-2px);
            color: white;
        }

        /* Estilo premium para los botones de descarga */
        div.stDownloadButton > button:first-child {
            background-color: #10b981; /* Verde esmeralda */
            color: white;
            border-radius: 8px;
            font-weight: 600;
            border: none;
            width: 100%;
            transition: all 0.3s ease;
        }
        div.stDownloadButton > button:first-child:hover {
            background-color: #059669;
            box-shadow: 0 10px 15px -3px rgba(16, 185, 129, 0.4);
            transform: translateY(-3px);
            color: white;
        }
        
        /* Ajuste de las tarjetas de subida de archivos */
        section[data-testid="stFileUploadDropzone"] {
            border-radius: 12px;
            border: 2px dashed #cbd5e1;
            background-color: #f8fafc;
            transition: all 0.3s;
        }
        section[data-testid="stFileUploadDropzone"]:hover {
            border-color: #3b82f6;
            background-color: #eff6ff;
        }
        </style>
    """, unsafe_allow_html=True)

def main():
    st.set_page_config(
        page_title="Agente CFS",
        page_icon="📦",
        layout="wide"
    )
    
    # Llamamos a los estilos mágicos aquí
    aplicar_estilos()
    
    # Inicializar session state
    if 'empresa_seleccionada' not in st.session_state:
        st.session_state.empresa_seleccionada = None
    if 'tipo_material' not in st.session_state:
        st.session_state.tipo_material = None
    if 'archivos_cargados' not in st.session_state:
        st.session_state.archivos_cargados = {}
    if 'session_id' not in st.session_state:
        st.session_state.session_id = str(hash(str(datetime.datetime.now())))
    
    st.title("📊 Agente CFS")
    st.markdown("---") # Una línea divisoria elegante
    
    # Mostrar pantalla inicial si no hay empresa seleccionada
    if st.session_state.empresa_seleccionada is None:
        mostrar_inicio_empresas()
    else:
        if st.session_state.tipo_material is None:
            if st.session_state.empresa_seleccionada == "Arauco":
                mostrar_menu_materiales_arauco()
            elif st.session_state.empresa_seleccionada == "CMPC":
                mostrar_menu_materiales_cmpc()
            elif st.session_state.empresa_seleccionada == "General":
                mostrar_menu_materiales_general()
        else:
            mostrar_panel_proceso()

def mostrar_inicio_empresas():
    st.header("Seleccione Empresa / Categoría")
    
    col1, col2, col3 = st.columns(3) # Cambiamos a 3 columnas
    
    with col1:
        if st.button("**ARAUCO**", use_container_width=True, type="primary"):
            st.session_state.empresa_seleccionada = "Arauco"
            st.session_state.tipo_material = None
            st.rerun()
    
    with col2:
        if st.button("**CMPC**", use_container_width=True, type="primary"):
            st.session_state.empresa_seleccionada = "CMPC"
            st.session_state.tipo_material = None
            st.rerun()

    with col3:
        if st.button("**GENERAL**", use_container_width=True, type="primary"):
            st.session_state.empresa_seleccionada = "General"
            st.session_state.tipo_material = None
            st.rerun()

# AGREGAR ESTA NUEVA FUNCIÓN
def mostrar_menu_materiales_general():
    st.header("General - Seleccione Proceso")
    
    if st.button("← Volver a Empresas"):
        st.session_state.empresa_seleccionada = None
        st.session_state.tipo_material = None
        st.rerun()
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("**Cuadrar Celulosa**", use_container_width=True):
            st.session_state.tipo_material = "Cuadrar Celulosa"
            st.rerun()
def mostrar_menu_materiales_arauco():
    st.header("Arauco - Seleccione Material")
    
    if st.button("← Volver a Empresas"):
        st.session_state.empresa_seleccionada = None
        st.session_state.tipo_material = None
        st.rerun()
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("**Celulosa**", use_container_width=True):
            st.session_state.tipo_material = "Celulosa"
            st.rerun()
        
        if st.button("**Madera**", use_container_width=True):
            st.session_state.tipo_material = "Madera"
            st.rerun()
    
    with col2:
        if st.button("**SAG**", use_container_width=True):
            st.session_state.tipo_material = "SAG"
            st.rerun()

def mostrar_menu_materiales_cmpc():
    st.header("CMPC - Seleccione Material")
    
    if st.button("← Volver a Empresas"):
        st.session_state.empresa_seleccionada = None
        st.session_state.tipo_material = None
        st.rerun()
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("**Celulosa**", use_container_width=True):
            st.session_state.tipo_material = "CMPC Celulosa"
            st.rerun()
        
        if st.button("**Papel**", use_container_width=True):
            st.session_state.tipo_material = "CMPC Papel"
            st.rerun()
    
    with col2:
        if st.button("**Madera**", use_container_width=True):
            st.session_state.tipo_material = "CMPC Madera"
            st.rerun()
        
        if st.button("**Plywood**", use_container_width=True):
            st.session_state.tipo_material = "CMPC Plywood"
            st.rerun()

def mostrar_panel_proceso():
    st.header(f"Panel: {st.session_state.tipo_material}")
    
    if st.button("← Volver"):
        st.session_state.tipo_material = None
        st.session_state.archivos_generados = None # Limpiar al volver
        st.rerun()
    
    lista_archivos = CONFIG_ARCHIVOS.get(st.session_state.tipo_material, [])
    
    # Limpiar archivos cargados si cambió el tipo de material
    if 'last_material' not in st.session_state or st.session_state.last_material != st.session_state.tipo_material:
        st.session_state.archivos_cargados = {}
        st.session_state.archivos_generados = None # Asegurarnos de limpiar salidas previas
        st.session_state.last_material = st.session_state.tipo_material
    
    st.subheader("Carga de Archivos")
    
# Crear formulario para subir archivos
    with st.form("upload_form"):
        for item in lista_archivos:
            es_multiple = item.get("multiple", False)
            required = "" if item["opcional"] else "🔴"
            
            # Obtener la descripción (si existe en la configuración)
            descripcion_tooltip = item.get("descripcion", None)
            
            if es_multiple:
                uploaded_files = st.file_uploader(
                    f"{required} {item['nombre']} {'(Múltiple)' if es_multiple else ''}",
                    type=['xlsx', 'xls', 'dbf'],
                    accept_multiple_files=True,
                    key=get_file_uploader_key(item["id"], st.session_state.session_id),
                    help=descripcion_tooltip  # <-- AQUÍ AGREGAMOS EL TOOLTIP
                )
                if uploaded_files:
                    temp_files = []
                    for uploaded_file in uploaded_files:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=f"_{uploaded_file.name}") as tmp_file:
                            tmp_file.write(uploaded_file.getvalue())
                            temp_files.append(tmp_file.name)
                    st.session_state.archivos_cargados[item["id"]] = temp_files
                    st.success(f"{len(uploaded_files)} archivo(s) cargado(s)")
            else:
                uploaded_file = st.file_uploader(
                    f"{required} {item['nombre']}",
                    type=['xlsx', 'xls', 'dbf'],
                    key=get_file_uploader_key(item["id"], st.session_state.session_id),
                    help=descripcion_tooltip  # <-- AQUÍ AGREGAMOS EL TOOLTIP
                )
                if uploaded_file:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=f"_{uploaded_file.name}") as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        st.session_state.archivos_cargados[item["id"]] = tmp_file.name
                    st.success(f"Archivo cargado: {uploaded_file.name}")
        
        submit_button = st.form_submit_button("🚀 Ejecutar Proceso")
    
    if submit_button:
        ejecutar_proceso()

    # --- AQUÍ ESTÁ LA MAGIA ---
    # Mostramos los botones FUERA del formulario y basados en session_state
    if st.session_state.get('archivos_generados'):
        st.subheader("📥 Archivos Generados")
        cols = st.columns(min(3, len(st.session_state.archivos_generados)))
        
        for idx, (nombre, archivo_bytes) in enumerate(st.session_state.archivos_generados):
            with cols[idx % 3]:
                st.download_button(
                    label=f"Descargar {nombre}",
                    data=archivo_bytes,
                    file_name=nombre,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"btn_descarga_{nombre}" # Es clave darle un ID único a cada botón
                )

import time # Asegúrate de tener 'import time' al inicio de tu app.py si no lo tienes

def ejecutar_proceso():
    # Validar archivos obligatorios
    lista_archivos = CONFIG_ARCHIVOS.get(st.session_state.tipo_material, [])
    faltantes = []
    
    for item in lista_archivos:
        if not item["opcional"] and item["id"] not in st.session_state.archivos_cargados:
            faltantes.append(item["nombre"])
    
    if faltantes:
        st.error(f"Faltan archivos obligatorios:\n- " + "\n- ".join(faltantes))
        return
    
    # Seleccionar lógica según tipo de material
    tipo_material = st.session_state.tipo_material
    rutas = st.session_state.archivos_cargados
    
    # ==========================================
    # PANTALLA DE CARGA ESTÉTICA
    # ==========================================
    with st.status("🚀 Iniciando procesamiento...", expanded=True) as status:
        st.write("📂 Leyendo archivos Excel...")
        time.sleep(0.3) # Pequeña pausa para que la animación se aprecie
        
        st.write("⚙️ Cruzando información y aplicando lógica de negocio...")
        
        # Aquí corre tu código pesado
        if tipo_material == "Madera":
            exito, mensaje, archivos = procesar_madera(rutas)
        elif tipo_material == "Celulosa":
            exito, mensaje, archivos = procesar_celulosa(rutas)
        elif tipo_material == "SAG":
            exito, mensaje, archivos = procesar_sag(rutas)
        elif tipo_material == "CMPC Celulosa":
            exito, mensaje, archivos = procesar_cmpc_celulosa(rutas)
        elif tipo_material == "CMPC Madera":
            exito, mensaje, archivos = procesar_cmpc_madera(rutas)
        elif tipo_material == "CMPC Papel":
            exito, mensaje, archivos = procesar_cmpc_papel(rutas)
        elif tipo_material == "CMPC Plywood":
            exito, mensaje, archivos = procesar_cmpc_plywood(rutas)
        elif tipo_material == "Cuadrar Celulosa": # <--- AGREGAR ESTAS DOS LÍNEAS
            exito, mensaje, archivos = procesar_cuadratura_celulosa(rutas)    
        else:
            exito, mensaje, archivos = False, "Lógica no implementada", []
        
        st.write("📝 Generando reportes de salida...")
        
        # Actualizamos el estado de la cajita dependiendo del resultado
        if exito:
            status.update(label="¡Procesamiento Completado!", state="complete", expanded=False)
        else:
            status.update(label="Ocurrió un error en el proceso", state="error", expanded=True)

    # Mostrar resultados y animaciones
    if exito:
        st.toast('¡Archivos generados con éxito!', icon='🎉') # Notificación flotante
        # st.balloons() # Descomenta esto si quieres globos volando por la pantalla (a veces es mucho, pero es divertido)
        
        st.success(mensaje)
        st.session_state.archivos_generados = archivos
    else:
        st.session_state.archivos_generados = None
        st.error(f"Error: {mensaje}")

if __name__ == "__main__":

    main()

